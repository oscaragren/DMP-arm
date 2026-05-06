from __future__ import annotations

import argparse
import json
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, Iterable, List, Optional, Sequence, Tuple

import numpy as np

from scipy.signal import butter, filtfilt
from typing import TYPE_CHECKING

from dmp.dmp import (
    DMPModel,
    canonical_phase,
    estimate_derivatives,
    learn_curvature_weights_from_demo,
    rollout_simple,
    rollout_simple_with_coupling,
    _solve_lwr_weights_multi,
)
from vis.plotting import plot_3d_trajectory, plot_angles_single, plot_dmp_single

if TYPE_CHECKING:
    import pandas as pd  # type: ignore


JOINT_LIMITS_DEG_NEW: np.ndarray = np.array(
    [
        [0.0, 60.0],  # 0: Elbow flexion
        [0.0, 80.0],  # 1: Shoulder flexion
        [5.0, 45.0],  # 2: Shoulder abduction
        [-60.0, 60.0],  # 3: Shoulder lateral/medial rotation
    ],
    dtype=float,
)


@dataclass(frozen=True)
class TrialId:
    subject: str
    motion: str
    trial: str

    @property
    def rel_dir(self) -> Path:
        return Path(self.subject, self.motion, self.trial)


def _json_dump(path: Path, obj: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with open(path, "w", encoding="utf-8") as f:
        json.dump(obj, f, indent=2, sort_keys=True)


def _ensure_dir(p: Path) -> None:
    p.mkdir(parents=True, exist_ok=True)


def _load_meta(trial_dir: Path) -> dict:
    meta_path = trial_dir / "meta.json"
    if meta_path.exists():
        with open(meta_path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def _discover_raw_trials(raw_root: Path) -> list[TrialId]:
    trials: list[TrialId] = []
    for meta_path in raw_root.glob("subject_*/*/trial_*/meta.json"):
        trial_dir = meta_path.parent
        motion_dir = trial_dir.parent
        subject_dir = motion_dir.parent
        trials.append(TrialId(subject=subject_dir.name, motion=motion_dir.name, trial=trial_dir.name))
    trials.sort(key=lambda x: (x.subject, x.motion, x.trial))
    return trials


def _parse_csv_or_all(values: Optional[str], *, kind: str) -> Optional[list[str]]:
    if values is None:
        return None
    v = values.strip()
    if v.lower() == "all":
        return None
    parts = [p.strip() for p in v.split(",") if p.strip()]
    if not parts:
        raise ValueError(f"--{kind} cannot be empty. Use e.g. 'all' or '1,2,3'.")
    return parts


def _filter_trials(
    all_trials: Sequence[TrialId],
    *,
    subjects: Optional[Sequence[str]],
    motions: Optional[Sequence[str]],
    trials: Optional[Sequence[str]],
) -> list[TrialId]:
    def _norm_subject(s: str) -> str:
        s = s.strip()
        if s.startswith("subject_"):
            return s
        if s.isdigit():
            return f"subject_{int(s):02d}"
        return s

    def _norm_trial(t: str) -> str:
        t = t.strip()
        if t.startswith("trial_"):
            return t
        if t.isdigit():
            return f"trial_{int(t):03d}"
        return t

    subjects_n = None if subjects is None else {_norm_subject(s) for s in subjects}
    motions_n = None if motions is None else {m.strip() for m in motions}
    trials_n = None if trials is None else {_norm_trial(t) for t in trials}

    out: list[TrialId] = []
    for tid in all_trials:
        if subjects_n is not None and tid.subject not in subjects_n:
            continue
        if motions_n is not None and tid.motion not in motions_n:
            continue
        if trials_n is not None and tid.trial not in trials_n:
            continue
        out.append(tid)
    return out


def _estimate_fps_from_t(t: np.ndarray, *, default_fps: float = 25.0) -> float:
    t = np.asarray(t, dtype=float).reshape(-1)
    if t.size < 2:
        return float(default_fps)
    dt = float(np.median(np.diff(t)))
    if not np.isfinite(dt) or dt <= 1e-6:
        return float(default_fps)
    return float(1.0 / dt)


def _interpolate_nans_1d(y: np.ndarray) -> np.ndarray:
    y = np.asarray(y, dtype=float).reshape(-1)
    if y.size == 0:
        return y.copy()
    x = np.arange(y.size, dtype=float)
    valid = np.isfinite(y)
    if int(np.sum(valid)) == 0:
        return y.copy()
    if int(np.sum(valid)) == 1:
        c = float(y[valid][0])
        return np.full_like(y, c, dtype=float)
    out = y.copy()
    out[~valid] = np.interp(x[~valid], x[valid], y[valid])
    return out


def interpolate_keypoints_cartesian(seq: np.ndarray) -> np.ndarray:
    """
    Interpolate NaNs over time for each keypoint coordinate.
    seq shape: (T, K, 3)
    """
    s = np.asarray(seq, dtype=float)
    if s.ndim != 3 or s.shape[2] != 3:
        raise ValueError(f"Expected seq shape (T,K,3), got {s.shape}")
    out = s.copy()
    T, K, D = out.shape
    for k in range(K):
        for d in range(D):
            out[:, k, d] = _interpolate_nans_1d(out[:, k, d])
    return out


def lowpass_keypoints(seq: np.ndarray, *, fps: float, cutoff_hz: float = 5.0, order: int = 2) -> np.ndarray:
    """
    Butterworth filtfilt per keypoint coordinate. seq shape (T,K,3), expected finite.
    """
    s = np.asarray(seq, dtype=float)
    if s.ndim != 3 or s.shape[2] != 3:
        raise ValueError(f"Expected seq shape (T,K,3), got {s.shape}")
    if s.shape[0] < 5:
        return s.copy()
    nyq = float(fps) * 0.5
    wn = min(float(cutoff_hz) / max(nyq, 1e-9), 0.99)
    b, a = butter(int(order), wn, btype="low")
    out = np.zeros_like(s, dtype=float)
    for k in range(s.shape[1]):
        for d in range(3):
            x = s[:, k, d]
            if not np.all(np.isfinite(x)):
                out[:, k, d] = x
                continue
            out[:, k, d] = filtfilt(b, a, x)
    return out


def sequence_to_angles_deg(seq: np.ndarray) -> np.ndarray:
    """
    Convert a (T, N>=6, 3) left-arm keypoint sequence to angles (T,4) in degrees.
    Angle order: [elbow_flex, shoulder_flex, shoulder_abd, shoulder_lat_med_rot]
    """
    from kinematics.simple_kinematics import get_angles as _get_angles_deg

    s = np.asarray(seq, dtype=float)
    if s.ndim != 3 or s.shape[2] != 3 or s.shape[1] < 6:
        raise ValueError(
            f"Expected seq shape (T, N>=6, 3) with "
            f"[left_shoulder, left_elbow, left_wrist, right_shoulder, left_hip, right_hip], got {s.shape}"
        )

    angles_deg = np.asarray(_get_angles_deg(s), dtype=float)  # (T,4) in degrees
    if angles_deg.ndim != 2 or angles_deg.shape[1] != 4:
        raise ValueError(f"Expected get_angles output shape (T,4), got {angles_deg.shape}")
    return angles_deg


def _clip_angles_deg(q_deg: np.ndarray) -> np.ndarray:
    q = np.asarray(q_deg, dtype=float).copy()
    if q.ndim != 2 or q.shape[1] != 4:
        raise ValueError(f"Expected q shape (T,4), got {q.shape}")
    return np.clip(q, JOINT_LIMITS_DEG_NEW[:, 0], JOINT_LIMITS_DEG_NEW[:, 1])


def _centers_and_widths(alpha_canonical: float, n_basis_functions: int) -> tuple[np.ndarray, np.ndarray]:
    centers = np.exp(-np.linspace(0, 1, int(n_basis_functions)) * float(alpha_canonical))
    if int(n_basis_functions) <= 1:
        widths = np.array([1.0], dtype=np.float64)
    else:
        d = np.diff(centers)
        d = np.hstack((d, [d[-1]]))
        widths = 1.0 / (d**2 + 1e-12)
    return centers.astype(np.float64), widths.astype(np.float64)


def fit_dmp_lwr_multi(
    demos: list[np.ndarray],
    *,
    tau: float,
    dt: float,
    n_basis_functions: int,
    alpha_canonical: float,
    alpha_transformation: float,
    beta_transformation: float,
    lwr_regularization: float = 1e-8,
) -> DMPModel:
    """
    Fit a DMP using `_solve_lwr_weights_multi` (as requested).

    demos: list of (T,4) in degrees. Must share same T.
    """
    if not demos:
        raise ValueError("demos must be non-empty")
    T = int(demos[0].shape[0])
    D = int(demos[0].shape[1])
    for i, q in enumerate(demos):
        if q.shape != (T, D):
            raise ValueError(f"demos[{i}] has shape {q.shape}, expected {(T, D)}")
    t_demo = np.arange(T, dtype=np.float64) * float(dt)
    x = canonical_phase(t_demo, tau=float(tau), alpha_canonical=float(alpha_canonical)).reshape(-1)

    centers, widths = _centers_and_widths(alpha_canonical, n_basis_functions)

    # Aggregate forcing targets across demos by concatenating time (keeps LWR assumptions simple).
    f_targets: list[np.ndarray] = []
    x_all: list[np.ndarray] = []
    for q in demos:
        q = np.asarray(q, dtype=float)
        q0 = q[0]
        g = q[-1]
        f_t = np.zeros((T, D), dtype=np.float64)
        for j in range(D):
            qj = q[:, j]
            dq, ddq = estimate_derivatives(
                qj,
                dt=float(dt),
                derivative_method="savgol",
                savgol_window_length=11,
                savgol_polyorder=3,
            )
            scale = float(g[j] - q0[j])
            if abs(scale) < 0.02:  # match repo's more stable scaling threshold
                scale = 1.0
            f_t[:, j] = (
                float(tau) ** 2 * ddq
                - float(alpha_transformation) * float(beta_transformation) * (float(g[j]) - qj)
                + float(alpha_transformation) * dq
            ) / scale
        f_targets.append(f_t)
        x_all.append(x)

    phase = np.concatenate(x_all, axis=0)
    f_target = np.concatenate(f_targets, axis=0)

    weights = _solve_lwr_weights_multi(
        phase=phase,
        f_target=f_target,
        centers=centers,
        widths=widths,
        regularization=float(lwr_regularization),
    )
    if weights.shape != (D, int(n_basis_functions)):
        raise ValueError(f"Expected weights shape {(D, int(n_basis_functions))}, got {weights.shape}")

    base_model = DMPModel(
        weights=np.asarray(weights, dtype=np.float64),
        centers=centers,
        widths=widths,
        alpha_canonical=float(alpha_canonical),
        alpha_transformation=float(alpha_transformation),
        beta_transformation=float(beta_transformation),
        tau=float(tau),
        n_joints=D,
        curvature_weights=np.zeros((D, int(n_basis_functions)), dtype=np.float64),
    )

    # Learn curvature weights per demo (trial-level), then mean.
    ridge_lambda = 1e-6
    curvature_weights_all = [
        learn_curvature_weights_from_demo(demo=np.asarray(d, dtype=float), model=base_model, dt=float(dt), ridge_lambda=ridge_lambda)
        for d in demos
    ]
    curvature_weights = np.mean(np.stack(curvature_weights_all, axis=0), axis=0)

    return DMPModel(
        weights=np.asarray(weights, dtype=np.float64),
        centers=centers,
        widths=widths,
        alpha_canonical=float(alpha_canonical),
        alpha_transformation=float(alpha_transformation),
        beta_transformation=float(beta_transformation),
        tau=float(tau),
        n_joints=D,
        curvature_weights=np.asarray(curvature_weights, dtype=np.float64),
    )


def save_angles_npz(out_dir: Path, *, angles_deg: np.ndarray, t: np.ndarray, meta: dict, dt: float) -> None:
    _ensure_dir(out_dir)
    np.savez(
        out_dir / "angles.npz",
        angles_deg=np.asarray(angles_deg, dtype=np.float64),
        t=np.asarray(t, dtype=np.float64),
        dt=float(dt),
        meta_json=json.dumps(meta),
    )


def save_dmp_model_npz(out_dir: Path, *, model: DMPModel) -> None:
    _ensure_dir(out_dir)
    np.savez(
        out_dir / "dmp_model.npz",
        weights=model.weights,
        centers=model.centers,
        widths=model.widths,
        alpha_canonical=model.alpha_canonical,
        alpha_transformation=model.alpha_transformation,
        beta_transformation=model.beta_transformation,
        tau=model.tau,
        n_joints=model.n_joints,
        curvature_weights=model.curvature_weights,
    )


def save_curvature_weights_npz(out_dir: Path, *, curvature_weights: np.ndarray, filename: str) -> None:
    _ensure_dir(out_dir)
    np.savez(out_dir / filename, curvature_weights=np.asarray(curvature_weights, dtype=np.float64))


def save_rollout_npz(
    out_dir: Path,
    *,
    q_gen_deg: np.ndarray,
    t: np.ndarray,
    dt: float,
    filename: str,
    q_gen_deg_clipped: Optional[np.ndarray] = None,
) -> None:
    _ensure_dir(out_dir)
    q = np.asarray(q_gen_deg, dtype=np.float64)
    q_clip = None if q_gen_deg_clipped is None else np.asarray(q_gen_deg_clipped, dtype=np.float64)
    np.savez(
        out_dir / filename,
        q_gen_deg=q,
        q_gen_deg_clipped=q_clip,
        t=np.asarray(t, dtype=np.float64),
        dt=float(dt),
        q0=q[0],
        qT=q[-1],
    )


def _rmse(a: np.ndarray, b: np.ndarray, *, axis: Optional[int | Tuple[int, ...]] = None) -> np.ndarray:
    a = np.asarray(a, dtype=float)
    b = np.asarray(b, dtype=float)
    if a.shape != b.shape:
        raise ValueError(f"RMSE needs same shape, got {a.shape} vs {b.shape}")
    return np.sqrt(np.mean((a - b) ** 2, axis=axis))


def _speed_magnitude(q: np.ndarray, *, dt: float) -> np.ndarray:
    q = np.asarray(q, dtype=float)
    if q.ndim == 1:
        dq = np.gradient(q, dt)
        return np.abs(dq)
    dq = np.gradient(q, dt, axis=0)
    return np.linalg.norm(dq, axis=1)


def _sparc_from_speed(
    speed: np.ndarray,
    *,
    fs: float,
    fc: float = 10.0,
    amp_thresh: float = 0.05,
    eps: float = 1e-12,
) -> float:
    v = np.asarray(speed, dtype=float).reshape(-1)
    n = int(v.size)
    if n < 8 or not np.all(np.isfinite(v)):
        return float("nan")
    v = v - np.mean(v)
    mag = np.abs(np.fft.rfft(v))
    freqs = np.fft.rfftfreq(n, d=1.0 / float(fs))
    mag_norm = mag / (np.max(mag) + eps)
    band = freqs <= float(fc)
    if not np.any(band):
        return float("nan")
    freqs_b = freqs[band]
    mag_b = mag_norm[band]
    keep = mag_b >= float(amp_thresh)
    if int(np.sum(keep)) < 2:
        return float("nan")
    freqs_k = freqs_b[keep]
    mag_k = mag_b[keep]
    A = np.log(mag_k + eps)
    df = np.diff(freqs_k)
    dA = np.diff(A)
    arc = float(np.sum(np.sqrt(df**2 + dA**2)))
    return -arc


def _ldlj(q: np.ndarray, *, dt: float, eps: float = 1e-12) -> float:
    x = np.asarray(q, dtype=float)
    if x.ndim == 1:
        x = x[:, None]
    if x.shape[0] < 8 or not np.all(np.isfinite(x)):
        return float("nan")
    duration = float((x.shape[0] - 1) * dt)
    if duration <= 0:
        return float("nan")
    A = float(np.linalg.norm(x[-1] - x[0]))
    A = max(A, eps)
    d1 = np.gradient(x, dt, axis=0)
    d2 = np.gradient(d1, dt, axis=0)
    d3 = np.gradient(d2, dt, axis=0)
    jerk_sq = np.sum(d3**2, axis=1)
    # Use trapz for broad NumPy compatibility (np.trapezoid is not present in older NumPy).
    integral = float(np.trapezoid(jerk_sq, dx=dt))
    dj = (duration**5 / (A**2)) * integral
    dj = max(dj, eps)
    return float(np.log(dj))


def evaluate_pair(q_demo: np.ndarray, q_gen: np.ndarray, *, dt: float) -> dict[str, Any]:
    qd = np.asarray(q_demo, dtype=float)
    qg = np.asarray(q_gen, dtype=float)
    if qd.shape != qg.shape:
        n = min(qd.shape[0], qg.shape[0])
        qd = qd[:n]
        qg = qg[:n]
    rmse_per = _rmse(qd, qg, axis=0).tolist()
    rmse_all = float(_rmse(qd, qg))
    fs = 1.0 / float(dt) if dt > 0 else float("nan")
    sparc_demo = _sparc_from_speed(_speed_magnitude(qd, dt=dt), fs=fs)
    sparc_gen = _sparc_from_speed(_speed_magnitude(qg, dt=dt), fs=fs)
    ldlj_demo = _ldlj(qd, dt=dt)
    ldlj_gen = _ldlj(qg, dt=dt)
    return {
        "rmse": {"overall": rmse_all, "per_dof": rmse_per},
        "sparc": {"demo": sparc_demo, "gen": sparc_gen, "delta": sparc_gen - sparc_demo},
        "ldlj": {"demo": ldlj_demo, "gen": ldlj_gen, "delta": ldlj_gen - ldlj_demo},
    }


def _trial_dirs_for_subject(processed_root: Path, subject: str) -> list[Path]:
    return sorted([p.parent for p in processed_root.glob(f"{subject}/*/trial_*/angles.npz")])


def compute_subject_mean_curvature_weights(processed_root: Path, subject: str) -> np.ndarray:
    weights: list[np.ndarray] = []
    for trial_dir in _trial_dirs_for_subject(processed_root, subject):
        cw_path = trial_dir / "curvature_weights_trial.npz"
        if not cw_path.exists():
            continue
        data = np.load(cw_path, allow_pickle=False)
        if "curvature_weights" not in data:
            continue
        w = np.asarray(data["curvature_weights"], dtype=float)
        if w.ndim == 2:
            weights.append(w)
    if not weights:
        raise FileNotFoundError(f"No trial curvature weights found for {subject} under {processed_root}")
    shapes = {tuple(w.shape) for w in weights}
    if len(shapes) != 1:
        raise ValueError(f"Curvature weight shapes differ for {subject}: {sorted(shapes)}")
    return np.mean(np.stack(weights, axis=0), axis=0)


def _subject_number(subject: str) -> int | None:
    """
    subject like 'subject_01' -> 1
    """
    if not isinstance(subject, str):
        return None
    if subject.startswith("subject_"):
        suf = subject.removeprefix("subject_")
        if suf.isdigit():
            return int(suf)
    return None


def _to_rows(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for r in records:
        tid = r["trial_id"]
        for variant in ["base", "personalized"]:
            m = r["metrics"][variant]
            row = {
                "subject": tid["subject"],
                "motion": tid["motion"],
                "trial": tid["trial"],
                "variant": variant,
                "rmse_overall": m["rmse"]["overall"],
                "rmse_dof0": m["rmse"]["per_dof"][0] if len(m["rmse"]["per_dof"]) > 0 else float("nan"),
                "rmse_dof1": m["rmse"]["per_dof"][1] if len(m["rmse"]["per_dof"]) > 1 else float("nan"),
                "rmse_dof2": m["rmse"]["per_dof"][2] if len(m["rmse"]["per_dof"]) > 2 else float("nan"),
                "rmse_dof3": m["rmse"]["per_dof"][3] if len(m["rmse"]["per_dof"]) > 3 else float("nan"),
                "sparc_demo": m["sparc"]["demo"],
                "sparc_gen": m["sparc"]["gen"],
                "sparc_delta": m["sparc"]["delta"],
                "ldlj_demo": m["ldlj"]["demo"],
                "ldlj_gen": m["ldlj"]["gen"],
                "ldlj_delta": m["ldlj"]["delta"],
            }
            rows.append(row)
    return rows


def _trial_num_from_id(trial_id: str) -> int | None:
    if isinstance(trial_id, str) and trial_id.startswith("trial_"):
        suf = trial_id.removeprefix("trial_")
        if suf.isdigit():
            return int(suf)
    return None


def _nanmean(a: Sequence[float]) -> float:
    x = np.asarray(list(a), dtype=float)
    if x.size == 0:
        return float("nan")
    return float(np.nanmean(x))


def _nanstd(a: Sequence[float]) -> float:
    x = np.asarray(list(a), dtype=float)
    n = int(np.sum(np.isfinite(x)))
    if n <= 1:
        return float("nan")
    return float(np.nanstd(x, ddof=1))


def _build_excel_tables(
    *,
    detailed_records: list[dict[str, Any]],
    processed_root: Path,
    created_at: str,
    source_json_name: str,
) -> dict[str, "pd.DataFrame"]:
    """
    Build a workbook layout similar to `results/quant_results_all_subjects.xlsx`.

    This returns a dict mapping sheet_name -> DataFrame.
    """
    import pandas as pd  # type: ignore

    # Build a per-trial table with both variants, then create derived sheets from it.
    rows: list[dict[str, Any]] = []
    for rec in detailed_records:
        tid = rec.get("trial_id") or {}
        subject = str(tid.get("subject", ""))
        motion = str(tid.get("motion", ""))
        trial = str(tid.get("trial", ""))
        trial_num = _trial_num_from_id(trial)
        trial_dir = Path(str(rec.get("trial_dir", "")))

        # Load dt/steps/dofs from angles.npz (keeps consistent with how we computed metrics)
        dt = float("nan")
        steps = float("nan")
        dofs = float("nan")
        try:
            with np.load(trial_dir / "angles.npz", allow_pickle=False) as a:
                q = np.asarray(a["angles_deg"], dtype=float)
                steps = int(q.shape[0])
                dofs = int(q.shape[1]) if q.ndim == 2 else 1
                dt = float(np.atleast_1d(a["dt"])[0]) if "dt" in a.files else float("nan")
        except Exception:
            pass

        m_base = (rec.get("metrics") or {}).get("base") or {}
        m_pers = (rec.get("metrics") or {}).get("personalized") or {}

        def _get(m: dict, path: tuple[str, ...], default=float("nan")) -> float:
            cur: Any = m
            for k in path:
                if not isinstance(cur, dict) or k not in cur:
                    return float(default)
                cur = cur[k]
            try:
                return float(cur)
            except Exception:
                return float(default)

        def _per_dof(m: dict, idx: int) -> float:
            try:
                arr = (m.get("rmse") or {}).get("per_dof") or []
                return float(arr[idx])
            except Exception:
                return float("nan")

        row = {
            "Subject": subject,
            "Trial": trial,
            "Trial #": trial_num,
            "Motion": motion,
            "dt": dt,
            "Steps": int(steps) if np.isfinite(steps) else np.nan,
            "DOFs": int(dofs) if np.isfinite(dofs) else np.nan,
            # Demo metrics duplicated for both variants (demo is the same)
            "LDLJ Demo": _get(m_base, ("ldlj", "demo")),
            "SPARC Demo": _get(m_base, ("sparc", "demo")),
            # Base
            "LDLJ Base": _get(m_base, ("ldlj", "gen")),
            "LDLJ Δ Base": _get(m_base, ("ldlj", "delta")),
            "SPARC Base": _get(m_base, ("sparc", "gen")),
            "SPARC Δ Base": _get(m_base, ("sparc", "delta")),
            "RMSE Overall Base": _get(m_base, ("rmse", "overall")),
            "RMSE DOF 1 Base": _per_dof(m_base, 0),
            "RMSE DOF 2 Base": _per_dof(m_base, 1),
            "RMSE DOF 3 Base": _per_dof(m_base, 2),
            "RMSE DOF 4 Base": _per_dof(m_base, 3),
            # Personalized
            "LDLJ Personalized": _get(m_pers, ("ldlj", "gen")),
            "LDLJ Δ Personalized": _get(m_pers, ("ldlj", "delta")),
            "SPARC Personalized": _get(m_pers, ("sparc", "gen")),
            "SPARC Δ Personalized": _get(m_pers, ("sparc", "delta")),
            "RMSE Overall Personalized": _get(m_pers, ("rmse", "overall")),
            "RMSE DOF 1 Personalized": _per_dof(m_pers, 0),
            "RMSE DOF 2 Personalized": _per_dof(m_pers, 1),
            "RMSE DOF 3 Personalized": _per_dof(m_pers, 2),
            "RMSE DOF 4 Personalized": _per_dof(m_pers, 3),
            "Trial Dir": str(trial_dir).replace(str(processed_root).rstrip("/") + "/", f"{processed_root.as_posix()}/"),
        }
        rows.append(row)

    df_all = pd.DataFrame(rows)
    df_all = df_all.sort_values(by=["Subject", "Trial #", "Trial"], na_position="last", kind="stable")

    # Overall Summary (2-column key/value style, like the reference workbook)
    overall_summary_rows = [
        ("Quantitative results workbook", ""),
        ("Created from", source_json_name),
        ("Created at", created_at),
        ("Trials selected", int(df_all.shape[0])),
    ]
    df_overall = pd.DataFrame(overall_summary_rows, columns=["Quantitative results workbook", "Unnamed: 1"])

    # Subject Summary (compute mean/std for each metric and variant)
    subjects = sorted(df_all["Subject"].dropna().unique().tolist())
    subj_rows: list[dict[str, Any]] = []
    for subj in subjects:
        sdf = df_all[df_all["Subject"] == subj]
        subj_rows.append(
            {
                "Subject": subj,
                "Trials": int(sdf.shape[0]),
                # LDLJ
                "LDLJ Demo Mean": _nanmean(sdf["LDLJ Demo"]),
                "LDLJ Demo Std": _nanstd(sdf["LDLJ Demo"]),
                "LDLJ Base Mean": _nanmean(sdf["LDLJ Base"]),
                "LDLJ Base Std": _nanstd(sdf["LDLJ Base"]),
                "LDLJ Δ Base Mean": _nanmean(sdf["LDLJ Δ Base"]),
                "LDLJ Δ Base Std": _nanstd(sdf["LDLJ Δ Base"]),
                "LDLJ Personalized Mean": _nanmean(sdf["LDLJ Personalized"]),
                "LDLJ Personalized Std": _nanstd(sdf["LDLJ Personalized"]),
                "LDLJ Δ Personalized Mean": _nanmean(sdf["LDLJ Δ Personalized"]),
                "LDLJ Δ Personalized Std": _nanstd(sdf["LDLJ Δ Personalized"]),
                # SPARC
                "SPARC Demo Mean": _nanmean(sdf["SPARC Demo"]),
                "SPARC Demo Std": _nanstd(sdf["SPARC Demo"]),
                "SPARC Base Mean": _nanmean(sdf["SPARC Base"]),
                "SPARC Base Std": _nanstd(sdf["SPARC Base"]),
                "SPARC Δ Base Mean": _nanmean(sdf["SPARC Δ Base"]),
                "SPARC Δ Base Std": _nanstd(sdf["SPARC Δ Base"]),
                "SPARC Personalized Mean": _nanmean(sdf["SPARC Personalized"]),
                "SPARC Personalized Std": _nanstd(sdf["SPARC Personalized"]),
                "SPARC Δ Personalized Mean": _nanmean(sdf["SPARC Δ Personalized"]),
                "SPARC Δ Personalized Std": _nanstd(sdf["SPARC Δ Personalized"]),
                # RMSE
                "RMSE Base Mean": _nanmean(sdf["RMSE Overall Base"]),
                "RMSE Base Std": _nanstd(sdf["RMSE Overall Base"]),
                "RMSE Personalized Mean": _nanmean(sdf["RMSE Overall Personalized"]),
                "RMSE Personalized Std": _nanstd(sdf["RMSE Overall Personalized"]),
            }
        )
    df_subject_summary = pd.DataFrame(subj_rows)

    # Plot data sheets (reference has 6 columns)
    df_ldlj_plot = df_all[["Subject", "Trial #", "Trial", "LDLJ Demo", "LDLJ Base", "LDLJ Δ Base"]].rename(
        columns={"LDLJ Base": "Base", "LDLJ Δ Base": "Delta", "LDLJ Demo": "Demo"}
    )
    df_ldlj_plot_p = df_all[["Subject", "Trial #", "Trial", "LDLJ Demo", "LDLJ Personalized", "LDLJ Δ Personalized"]].rename(
        columns={"LDLJ Personalized": "Personalized", "LDLJ Δ Personalized": "Delta", "LDLJ Demo": "Demo"}
    )
    df_sparc_plot = df_all[["Subject", "Trial #", "Trial", "SPARC Demo", "SPARC Base", "SPARC Δ Base"]].rename(
        columns={"SPARC Base": "Base", "SPARC Δ Base": "Delta", "SPARC Demo": "Demo"}
    )
    df_sparc_plot_p = df_all[["Subject", "Trial #", "Trial", "SPARC Demo", "SPARC Personalized", "SPARC Δ Personalized"]].rename(
        columns={"SPARC Personalized": "Personalized", "SPARC Δ Personalized": "Delta", "SPARC Demo": "Demo"}
    )

    # Per-subject sheets (reference shows trial-level LDLJ/SPARC/RMSE). We'll include both variants.
    per_subject_sheets: dict[str, pd.DataFrame] = {}
    for subj in subjects:
        sdf = df_all[df_all["Subject"] == subj].copy()
        sdf = sdf.sort_values(by=["Trial #", "Trial"], na_position="last", kind="stable")
        per_subject_sheets[subj] = sdf[
            [
                "Trial #",
                "Trial",
                "LDLJ Demo",
                "LDLJ Base",
                "LDLJ Δ Base",
                "LDLJ Personalized",
                "LDLJ Δ Personalized",
                "SPARC Demo",
                "SPARC Base",
                "SPARC Δ Base",
                "SPARC Personalized",
                "SPARC Δ Personalized",
                "RMSE Overall Base",
                "RMSE Overall Personalized",
            ]
        ]

    sheets: dict[str, pd.DataFrame] = {
        "Overall Summary": df_overall,
        "Subject Summary": df_subject_summary,
        "All Trials": df_all,
        "LDLJ Plot Data (Base)": df_ldlj_plot,
        "LDLJ Plot Data (Personalized)": df_ldlj_plot_p,
        "SPARC Plot Data (Base)": df_sparc_plot,
        "SPARC Plot Data (Personalized)": df_sparc_plot_p,
    }
    sheets.update(per_subject_sheets)
    return sheets


def _autosize_openpyxl_columns(xlsx_path: Path) -> None:
    try:
        from openpyxl import load_workbook  # type: ignore
        from openpyxl.utils import get_column_letter  # type: ignore
    except Exception:
        return
    wb = load_workbook(xlsx_path)
    for ws in wb.worksheets:
        # Freeze header row for table-like sheets
        if ws.max_row >= 2 and ws.max_column >= 2:
            ws.freeze_panes = "A2"
        for col in range(1, ws.max_column + 1):
            letter = get_column_letter(col)
            max_len = 0
            # cap to avoid huge width on 'Trial Dir'
            for row in range(1, min(ws.max_row, 200) + 1):
                v = ws.cell(row=row, column=col).value
                if v is None:
                    continue
                s = str(v)
                if len(s) > max_len:
                    max_len = len(s)
            ws.column_dimensions[letter].width = min(max(10, max_len + 2), 60)
    wb.save(xlsx_path)


def main() -> None:
    ap = argparse.ArgumentParser(description="End-to-end quantitative analysis over data/raw.")
    ap.add_argument("--raw-root", type=Path, default=Path("data/raw"))
    ap.add_argument("--processed-root", type=Path, default=Path("data/processed"))
    ap.add_argument("--results-root", type=Path, default=Path("results"))

    ap.add_argument("--subjects", type=str, default="all")
    ap.add_argument("--motions", type=str, default="all")
    ap.add_argument("--trials", type=str, default="all")

    ap.add_argument("--cutoff-hz", type=float, default=20.0)
    ap.add_argument("--filter-order", type=int, default=2)

    ap.add_argument("--n-basis", type=int, default=100)
    ap.add_argument("--tau", type=float, default=1.0)
    ap.add_argument("--alpha-canonical", type=float, default=4.0)
    ap.add_argument("--alpha-transformation", type=float, default=25.0)
    ap.add_argument("--beta-transformation", type=float, default=6.25)
    args = ap.parse_args()

    raw_root = args.raw_root
    processed_root = args.processed_root
    results_root = args.results_root

    subjects = _parse_csv_or_all(args.subjects, kind="subjects")
    motions = _parse_csv_or_all(args.motions, kind="motions")
    trials = _parse_csv_or_all(args.trials, kind="trials")

    all_trials = _discover_raw_trials(raw_root)
    selected = _filter_trials(all_trials, subjects=subjects, motions=motions, trials=trials)
    if not selected:
        raise FileNotFoundError(f"No raw trials found under {raw_root} for selection.")

    # -------------------------
    # Pass 1: per-trial processing + DMP fit + base rollout + save plots
    # -------------------------
    failures: list[dict[str, str]] = []
    per_trial: list[dict[str, Any]] = []

    for tid in selected:
        raw_trial_dir = raw_root / tid.rel_dir
        out_dir = processed_root / tid.rel_dir
        _ensure_dir(out_dir)
        meta = _load_meta(raw_trial_dir)

        try:
            seq = np.load(raw_trial_dir / "left_arm_seq_camera.npy")
            t = np.load(raw_trial_dir / "left_arm_t.npy")
            if t.ndim != 1 or t.shape[0] != seq.shape[0]:
                t = np.arange(seq.shape[0], dtype=np.float64)

            # 2) Interpolate in cartesian space (keypoints)
            seq_i = interpolate_keypoints_cartesian(seq)

            # 3) Low-pass filter in cartesian space
            fps = _estimate_fps_from_t(t, default_fps=float(meta.get("fps_nominal", 25.0)))
            seq_f = lowpass_keypoints(seq_i, fps=fps, cutoff_hz=float(args.cutoff_hz), order=int(args.filter_order))

            # 4) Convert to angles (degrees), clip to new limits
            q_demo = sequence_to_angles_deg(seq_f)
            valid = np.all(np.isfinite(q_demo), axis=1)
            q_demo = q_demo[valid]
            t_demo = t[valid]
            if q_demo.shape[0] < 10:
                raise ValueError(f"Not enough valid angle samples after interpolation/filter: {q_demo.shape}")

            # Uniform dt for DMP fit/rollouts (normalized to tau)
            dt = float(args.tau) / float(q_demo.shape[0] - 1)

            # 5) Fit DMP (LWR multi)
            model_trial = fit_dmp_lwr_multi(
                [q_demo],
                tau=float(args.tau),
                dt=dt,
                n_basis_functions=int(args.n_basis),
                alpha_canonical=float(args.alpha_canonical),
                alpha_transformation=float(args.alpha_transformation),
                beta_transformation=float(args.beta_transformation),
            )

            # Save per-trial curvature weights for later subject averaging
            save_curvature_weights_npz(out_dir, curvature_weights=model_trial.curvature_weights, filename="curvature_weights_trial.npz")

            # 6) Rollout without personalization (no coupling)
            model_base = DMPModel(
                weights=model_trial.weights,
                centers=model_trial.centers,
                widths=model_trial.widths,
                alpha_canonical=model_trial.alpha_canonical,
                alpha_transformation=model_trial.alpha_transformation,
                beta_transformation=model_trial.beta_transformation,
                tau=model_trial.tau,
                n_joints=model_trial.n_joints,
                curvature_weights=np.zeros_like(model_trial.curvature_weights),
            )
            # Keep an unclipped rollout for evaluation; keep a clipped copy for saving/plotting.
            q_gen_base_unclipped = rollout_simple(model_base, q_demo[0], q_demo[-1], tau=float(args.tau), dt=dt)
            q_gen_base_clipped = _clip_angles_deg(q_gen_base_unclipped)

            # 7) Save angles, model, rollout
            save_angles_npz(out_dir, angles_deg=q_demo, t=t_demo, meta=meta, dt=dt)
            save_dmp_model_npz(out_dir, model=model_trial)
            save_rollout_npz(
                out_dir,
                q_gen_deg=q_gen_base_unclipped,
                q_gen_deg_clipped=q_gen_base_clipped,
                t=t_demo,
                dt=dt,
                filename="dmp_rollout_base.npz",
            )

            # 9) Plots
            plot_3d_trajectory(seq_f, t, meta, out_dir / "trajectory_3d.png")
            plot_angles_single(
                elbow_rad=np.deg2rad(q_demo[:, 0]),
                shoulder_rad=np.deg2rad(q_demo[:, 1:4]),
                t=t_demo,
                meta=meta,
                title_suffix="angles",
                units="deg",
                out_path=out_dir / "angles.png",
            )
            plot_dmp_single(
                q_demo=q_demo,
                q_gen=q_gen_base_unclipped,
                meta=meta,
                title_suffix="dmp_base",
                out_path=out_dir / "dmp_base.png",
            )

            per_trial.append(
                {
                    "trial_id": asdict(tid),
                    "raw_trial_dir": str(raw_trial_dir),
                    "processed_trial_dir": str(out_dir),
                    "n_steps": int(q_demo.shape[0]),
                }
            )
        except Exception as e:
            failures.append({"trial_id": json.dumps(asdict(tid)), "error": f"{type(e).__name__}: {e}"})

    # -------------------------
    # Pass 2: per-subject mean curvature weights + personalized rollouts + metrics
    # -------------------------
    subjects_seen = sorted({r["trial_id"]["subject"] for r in per_trial})
    subject_mean_paths: dict[str, str] = {}
    subject_mean_weights: dict[str, np.ndarray] = {}
    coupling_dir = Path("coupling")
    _ensure_dir(coupling_dir)
    for subj in subjects_seen:
        try:
            w_mean = compute_subject_mean_curvature_weights(processed_root, subj)
            subject_mean_weights[subj] = w_mean
            subj_dir = processed_root / subj
            _ensure_dir(subj_dir)
            out_path = subj_dir / "curvature_weights_mean.npz"
            np.savez(out_path, curvature_weights=w_mean, n_trials=int(len(_trial_dirs_for_subject(processed_root, subj))))
            subject_mean_paths[subj] = str(out_path)

            # Also save in coupling/ with thesis naming convention, e.g. S01_curv_weights_mean.npz
            s_num = _subject_number(subj)
            if s_num is not None:
                out_coupling = coupling_dir / f"S{s_num:02d}_curv_weights_mean.npz"
                np.savez(out_coupling, curvature_weights=w_mean, subject=subj)
        except Exception as e:
            failures.append({"trial_id": json.dumps({"subject": subj}), "error": f"SubjectMeanError: {type(e).__name__}: {e}"})

    # Evaluate + write personalized rollouts per trial
    detailed_records: list[dict[str, Any]] = []
    for r in per_trial:
        tid = TrialId(**r["trial_id"])
        trial_dir = processed_root / tid.rel_dir
        try:
            angles_npz = np.load(trial_dir / "angles.npz", allow_pickle=False)
            q_demo = np.asarray(angles_npz["angles_deg"], dtype=float)
            dt = float(np.atleast_1d(angles_npz["dt"])[0]) if "dt" in angles_npz.files else float(args.tau) / (q_demo.shape[0] - 1)

            base_npz = np.load(trial_dir / "dmp_rollout_base.npz", allow_pickle=False)
            q_base = np.asarray(base_npz["q_gen_deg"], dtype=float)  # unclipped rollout (preferred for metrics)
            q_base_plot = (
                np.asarray(base_npz["q_gen_deg_clipped"], dtype=float)
                if "q_gen_deg_clipped" in base_npz.files and base_npz["q_gen_deg_clipped"] is not None
                else q_base
            )

            model_npz = np.load(trial_dir / "dmp_model.npz", allow_pickle=False)
            model = DMPModel(
                weights=np.asarray(model_npz["weights"], dtype=float),
                centers=np.asarray(model_npz["centers"], dtype=float),
                widths=np.asarray(model_npz["widths"], dtype=float),
                alpha_canonical=float(np.atleast_1d(model_npz["alpha_canonical"])[0]),
                alpha_transformation=float(np.atleast_1d(model_npz["alpha_transformation"])[0]),
                beta_transformation=float(np.atleast_1d(model_npz["beta_transformation"])[0]),
                tau=float(np.atleast_1d(model_npz["tau"])[0]),
                n_joints=int(np.atleast_1d(model_npz["n_joints"])[0]),
                curvature_weights=np.asarray(model_npz["curvature_weights"], dtype=float),
            )

            subj = tid.subject
            if subj not in subject_mean_weights:
                raise KeyError(f"No subject mean curvature weights for {subj}")
            save_curvature_weights_npz(
                trial_dir,
                curvature_weights=np.asarray(subject_mean_weights[subj], dtype=float),
                filename="curvature_weights_subject_mean.npz",
            )
            model_personalized = DMPModel(
                weights=model.weights,
                centers=model.centers,
                widths=model.widths,
                alpha_canonical=model.alpha_canonical,
                alpha_transformation=model.alpha_transformation,
                beta_transformation=model.beta_transformation,
                tau=model.tau,
                n_joints=model.n_joints,
                curvature_weights=np.asarray(subject_mean_weights[subj], dtype=float),
            )

            q_personalized = rollout_simple_with_coupling(
                model_personalized, q_demo[0], q_demo[-1], tau=float(args.tau), dt=dt
            )
            q_personalized_unclipped = np.asarray(q_personalized, dtype=float)
            q_personalized_clipped = _clip_angles_deg(q_personalized_unclipped)
            save_rollout_npz(
                trial_dir,
                q_gen_deg=q_personalized_unclipped,
                q_gen_deg_clipped=q_personalized_clipped,
                t=np.asarray(angles_npz["t"], dtype=float),
                dt=dt,
                filename="dmp_rollout_personalized.npz",
            )

            # Plots
            meta = _load_meta(Path(r["raw_trial_dir"]))
            plot_dmp_single(
                q_demo=q_demo,
                q_gen=q_personalized_unclipped,
                meta=meta,
                title_suffix="dmp_personalized",
                out_path=trial_dir / "dmp_personalized.png",
            )

            # Also save a plot where the *clipped* base + personalized DMPs are shown together.
            # This is useful for visualizing what would be sent to the robot/controller.
            def _plot_dmp_clipped_base_vs_personalized(
                *,
                q_demo_deg: np.ndarray,
                q_base_clipped_deg: np.ndarray,
                q_personalized_clipped_deg: np.ndarray,
                meta: dict,
                out_path: Path,
            ) -> None:
                import matplotlib.pyplot as plt

                joint_names = [
                    "Elbow flexion",
                    "Shoulder flexion",
                    "Shoulder abduction",
                    "Shoulder internal rotation",
                ]

                tau = 1.0
                t_demo = np.linspace(0.0, tau, q_demo_deg.shape[0])
                t_base = np.linspace(0.0, tau, q_base_clipped_deg.shape[0])
                t_pers = np.linspace(0.0, tau, q_personalized_clipped_deg.shape[0])

                fig, axes = plt.subplots(2, 2, figsize=(10, 8), sharex=True)
                axes = axes.flatten()
                for j in range(4):
                    ax = axes[j]
                    ax.plot(t_demo, q_demo_deg[:, j], color="#3498db", linewidth=1.5, label="Demo")
                    ax.plot(t_base, q_base_clipped_deg[:, j], color="#F58518", linestyle="--", linewidth=1.2, label="Base (clipped)")
                    ax.plot(
                        t_pers,
                        q_personalized_clipped_deg[:, j],
                        color="#54A24B",
                        linestyle="--",
                        linewidth=1.2,
                        label="Personalized (clipped)",
                    )
                    ax.set_title(joint_names[j])
                    ax.set_ylabel("Angle (deg)")
                    ax.grid(True, alpha=0.3)
                    ax.legend(loc="upper right", fontsize=8)
                axes[-1].set_xlabel("Time (normalized)")

                subject = meta.get("subject", "?")
                motion = meta.get("motion", "?")
                trial = meta.get("trial", "?")
                fig.suptitle(
                    f"DMP trajectories (clipped) — subject {subject}, {motion}, trial {trial}",
                    fontsize=11,
                )
                plt.tight_layout()
                out_path.parent.mkdir(parents=True, exist_ok=True)
                plt.savefig(out_path, dpi=140)
                plt.close(fig)

            # Prefer base clipped from saved artifact when present (matches what gets exported).
            q_base_clipped_for_plot = np.asarray(q_base_plot, dtype=float)
            _plot_dmp_clipped_base_vs_personalized(
                q_demo_deg=q_demo,
                q_base_clipped_deg=q_base_clipped_for_plot,
                q_personalized_clipped_deg=q_personalized_clipped,
                meta=meta,
                out_path=trial_dir / "dmp_clipped_base_vs_personalized.png",
            )

            # Metrics should be computed on the *unclipped* DMP outputs (as requested).
            metrics_base = evaluate_pair(q_demo, q_base, dt=dt)
            metrics_pers = evaluate_pair(q_demo, q_personalized_unclipped, dt=dt)

            record = {
                "trial_id": asdict(tid),
                "trial_dir": str(trial_dir),
                "metrics": {"base": metrics_base, "personalized": metrics_pers},
                "artifacts": {
                    "rollout_base": str((trial_dir / "dmp_rollout_base.npz").resolve()),
                    "rollout_personalized": str((trial_dir / "dmp_rollout_personalized.npz").resolve()),
                    "curvature_trial": str((trial_dir / "curvature_weights_trial.npz").resolve()),
                    "curvature_subject_mean": str((trial_dir / "curvature_weights_subject_mean.npz").resolve()),
                },
            }
            detailed_records.append(record)

            _json_dump(trial_dir / "quant_results.json", record)
        except Exception as e:
            failures.append({"trial_id": json.dumps(asdict(tid)), "error": f"{type(e).__name__}: {e}"})

    # -------------------------
    # Aggregation outputs (JSON + Excel)
    # -------------------------
    results_root.mkdir(parents=True, exist_ok=True)
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    overall = {
        "created_at": datetime.now().isoformat(timespec="seconds"),
        "raw_root": str(raw_root),
        "processed_root": str(processed_root),
        "selection": {
            "subjects": subjects if subjects is not None else "all",
            "motions": motions if motions is not None else "all",
            "trials": trials if trials is not None else "all",
        },
        "hyperparams": {
            "cutoff_hz": float(args.cutoff_hz),
            "filter_order": int(args.filter_order),
            "n_basis": int(args.n_basis),
            "tau": float(args.tau),
            "alpha_canonical": float(args.alpha_canonical),
            "alpha_transformation": float(args.alpha_transformation),
            "beta_transformation": float(args.beta_transformation),
        },
        "subject_curvature_weights_mean": subject_mean_paths,
        "n_raw_trials_found": int(len(all_trials)),
        "n_trials_selected": int(len(selected)),
        "n_trials_processed": int(len(per_trial)),
        "n_trials_evaluated": int(len(detailed_records)),
        "trials": detailed_records,
        "failures": failures,
    }

    out_json = results_root / f"quant_analysis_{stamp}.json"
    _json_dump(out_json, overall)

    # Excel (reference-like workbook layout)
    out_xlsx = results_root / f"quant_analysis_{stamp}.xlsx"
    try:
        import pandas as pd  # type: ignore

        # Prefer openpyxl (common), otherwise try xlsxwriter.
        try:
            with pd.ExcelWriter(out_xlsx, engine="openpyxl") as writer:
                sheets = _build_excel_tables(
                    detailed_records=detailed_records,
                    processed_root=processed_root,
                    created_at=str(overall.get("created_at", "")),
                    source_json_name=out_json.name,
                )
                for name, df_sheet in sheets.items():
                    safe = str(name)[:31]  # Excel sheet name limit
                    df_sheet.to_excel(writer, index=False, sheet_name=safe)
        except ModuleNotFoundError:
            with pd.ExcelWriter(out_xlsx, engine="xlsxwriter") as writer:
                sheets = _build_excel_tables(
                    detailed_records=detailed_records,
                    processed_root=processed_root,
                    created_at=str(overall.get("created_at", "")),
                    source_json_name=out_json.name,
                )
                for name, df_sheet in sheets.items():
                    safe = str(name)[:31]
                    df_sheet.to_excel(writer, index=False, sheet_name=safe)

        # Best-effort: autosize columns and freeze headers (openpyxl only)
        _autosize_openpyxl_columns(out_xlsx)
    except Exception as e:
        failures.append({"trial_id": "excel", "error": f"{type(e).__name__}: {e}"})
        overall["failures"] = failures
        _json_dump(out_json, overall)

    print(
        json.dumps(
            {
                "processed_trials": len(per_trial),
                "evaluated_trials": len(detailed_records),
                "failures": len(failures),
                "results_json": str(out_json.resolve()),
                "results_xlsx": str(out_xlsx.resolve()),
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()

